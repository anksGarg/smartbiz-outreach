from __future__ import annotations

import base64
import io
import json
import os
import re
import secrets
import struct
import uuid
from concurrent.futures import ThreadPoolExecutor
from datetime import date, datetime, timedelta, timezone
from zoneinfo import ZoneInfo

_TZ = ZoneInfo("America/Los_Angeles")
from pathlib import Path

import anthropic
import gspread
import httpx
import pandas as pd
from dotenv import load_dotenv
from google.oauth2.service_account import Credentials as _SACredentials
import qrcode as _qrlib
from fastapi import FastAPI, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, Response, StreamingResponse
from pyproj import Transformer
from pydantic import BaseModel

load_dotenv()

app = FastAPI(title="SmallBizOutreach API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "https://smartbiz-outreach-1.onrender.com",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

DATA_DIR = Path(__file__).parent / "data"
CONTACTS_FILE = DATA_DIR / "contacts.json"
CAMPAIGNS_FILE = DATA_DIR / "campaigns.json"

# Test mode: all emails go here instead of the real contact email
TEST_RECIPIENT = "ajaygarg12@gmail.com"

# Column name aliases — maps whatever the CSV/XLSX header says to our schema key
COLUMN_ALIASES = {
    "name": ["name", "full name", "fullname", "contact name", "contact", "decision-maker / contact", "decision maker / contact", "decision-maker/contact", "property manager"],
    "company": ["company", "company name", "business", "organization", "org", "center name", "property name", "site name", "business / institution name", "business/institution name", "institution name"],
    "phone": ["phone", "phone number", "tel", "telephone", "mobile", "cell"],
    "email": ["email", "email address", "e-mail"],
    "address": ["address", "street address", "mailing address", "location", "street address / area", "street address/area"],
    "property_type": ["property type", "property category", "contact type"],
    "priority_tier": ["priority tier", "priority"],
}


def _normalize_priority(val: str) -> str:
    v = val.strip().upper()
    if v in ("A", "1") or v.startswith("TIER 1"):
        return "Tier 1"
    if v in ("B", "2") or v.startswith("TIER 2"):
        return "Tier 2"
    if v in ("C", "3") or v.startswith("TIER 3"):
        return "Tier 3"
    return val.strip()


def _load_campaigns() -> list[dict]:
    if not CAMPAIGNS_FILE.exists():
        return []
    try:
        return json.loads(CAMPAIGNS_FILE.read_text())
    except (json.JSONDecodeError, OSError):
        return []


def _save_campaigns(campaigns: list[dict]) -> None:
    CAMPAIGNS_FILE.write_text(json.dumps(campaigns, indent=2))


def _load_contacts() -> list[dict]:
    if not CONTACTS_FILE.exists():
        return []
    try:
        return json.loads(CONTACTS_FILE.read_text())
    except (json.JSONDecodeError, OSError):
        return []


def _save_contacts(contacts: list[dict]) -> None:
    CONTACTS_FILE.write_text(json.dumps(contacts, indent=2))


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Rename incoming columns to our canonical schema keys."""
    rename_map = {}
    lowered = {col.strip().lower(): col for col in df.columns}
    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in lowered:
                rename_map[lowered[alias]] = field
                break
    return df.rename(columns=rename_map)


def _parse_file(file: UploadFile) -> pd.DataFrame:
    content = file.file.read()
    filename = (file.filename or "").lower()
    try:
        if filename.endswith(".xlsx"):
            return pd.read_excel(io.BytesIO(content), engine="openpyxl")
        elif filename.endswith(".csv"):
            return pd.read_csv(io.StringIO(content.decode("utf-8", errors="replace")))
        else:
            raise HTTPException(status_code=400, detail="Only .csv and .xlsx files are supported.")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not parse file: {exc}")


@app.post("/contacts/preview")
def preview_contacts(file: UploadFile):
    df = _parse_file(file)
    df = _normalize_columns(df)

    schema_fields = list(COLUMN_ALIASES.keys())
    for field in schema_fields:
        if field not in df.columns:
            df[field] = ""

    df = df[schema_fields].copy()
    df = df.where(pd.notna(df), "")
    df = df.astype(str)

    rows = df.head(100).to_dict(orient="records")
    return {"rows": rows, "total": len(df)}


@app.get("/contacts")
def list_contacts():
    return _load_contacts()


@app.get("/health")
def health_check():
    return {"status": "ok"}


@app.get("/favicon.ico", include_in_schema=False)
def favicon():
    svg = (
        '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 32 32">'
        '<rect width="32" height="32" rx="4" fill="#1e3a5f"/>'
        '<polygon points="16,5 29,16 25,16 25,27 19,27 19,20 13,20 13,27 7,27 7,16 3,16" fill="#ffffff"/>'
        '</svg>'
    )
    return Response(content=svg.encode(), media_type="image/svg+xml",
                    headers={"Cache-Control": "public, max-age=86400"})


@app.post("/contacts/import")
def import_contacts(file: UploadFile):
    df = _parse_file(file)
    df = _normalize_columns(df)

    schema_fields = list(COLUMN_ALIASES.keys())
    for field in schema_fields:
        if field not in df.columns:
            df[field] = ""

    df = df[schema_fields].copy()
    df = df.where(pd.notna(df), "")
    df = df.astype(str)
    df["email"] = df["email"].str.strip().str.lower()
    df["priority_tier"] = df["priority_tier"].apply(lambda v: _normalize_priority(v) if v else "")

    existing = _load_contacts()
    existing_emails = {c.get("email", "").strip().lower() for c in existing if c.get("email")}

    imported, skipped = 0, 0
    for row in df.to_dict(orient="records"):
        email = row.get("email", "")
        if email and email in existing_emails:
            skipped += 1
            continue
        existing.append({**row, "last_contacted": None})
        if email:
            existing_emails.add(email)
        imported += 1

    _save_contacts(existing)
    return {"imported": imported, "skipped": skipped}


class DraftRequest(BaseModel):
    name: str = ""
    company: str = ""
    phone: str = ""
    email: str = ""
    address: str = ""
    property_type: str = ""


_TITLES = {"mr", "mrs", "ms", "miss", "dr", "prof", "rev", "sir"}


def _first_name(full_name: str) -> str:
    """Extract first name from a raw name field that may include title, role, or other details."""
    # Drop everything after the first comma (e.g. ", Senior Property Manager – Retail")
    base = full_name.split(",")[0].strip()
    words = base.split()
    # Skip leading title tokens (Mr. / Dr. etc.)
    for word in words:
        if word.rstrip(".").lower() not in _TITLES:
            return word
    return words[0] if words else full_name


@app.post("/draft")
def draft_message(req: DraftRequest):
    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="ANTHROPIC_API_KEY not configured.")

    first = _first_name(req.name) if req.name else ""
    parts = []
    if first:
        parts.append(f"Contact first name: {first}")
    if req.company:
        parts.append(f"Company: {req.company}")
    if req.address:
        parts.append(f"Address: {req.address}")
    if req.property_type:
        parts.append(f"Property type: {req.property_type}")
    contact_info = "\n".join(parts) if parts else "No contact details provided."

    try:
        client = anthropic.Anthropic(api_key=api_key)
        response = client.messages.create(
            model="claude-opus-4-8",
            max_tokens=300,
            system=(
                "You are writing outreach messages for a roofing and gutter company in Portland, OR. "
                "Write short, professional, personalized messages. "
                "Keep it under 100 words. Never sound like a mass email."
            ),
            messages=[
                {
                    "role": "user",
                    "content": f"Write an outreach message for this contact:\n{contact_info}",
                }
            ],
        )
        message = response.content[0].text
        return {"message": message}
    except anthropic.APIError as e:
        raise HTTPException(status_code=502, detail=f"Claude API error: {e}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Could not draft message: {e}")


class SendRequest(BaseModel):
    name: str = ""
    company: str = ""
    email: str = ""
    message: str = ""


@app.post("/send")
def send_email(req: SendRequest):
    brevo_key = os.getenv("BREVO_API_KEY")
    from_email = os.getenv("BREVO_FROM_EMAIL")
    from_name = os.getenv("BREVO_FROM_NAME", "Willamette Power Roofing")

    if not brevo_key or brevo_key == "your_brevo_api_key_here":
        raise HTTPException(status_code=500, detail="BREVO_API_KEY not configured in .env")
    if not from_email:
        raise HTTPException(status_code=500, detail="BREVO_FROM_EMAIL not configured in .env")
    if not req.message.strip():
        raise HTTPException(status_code=400, detail="Message cannot be empty.")

    subject = f"A note for {req.name}" if req.name else "Message from Willamette Power Roofing"

    payload = {
        "sender": {"name": from_name, "email": from_email},
        "to": [{"email": TEST_RECIPIENT, "name": req.name or "Test"}],
        "subject": subject,
        "textContent": req.message,
    }

    try:
        resp = httpx.post(
            "https://api.brevo.com/v3/smtp/email",
            headers={"api-key": brevo_key, "Content-Type": "application/json"},
            json=payload,
            timeout=15,
        )
        resp.raise_for_status()
    except httpx.HTTPStatusError as e:
        detail = e.response.text[:300] if e.response else str(e)
        raise HTTPException(status_code=502, detail=f"Brevo error: {detail}")
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Could not send email: {e}")

    # Log to campaigns.json
    campaigns = _load_campaigns()
    campaigns.append({
        "sent_at": datetime.now(timezone.utc).isoformat(),
        "contact_name": req.name,
        "contact_email": req.email,
        "sent_to": TEST_RECIPIENT,
        "message": req.message,
        "status": "sent",
    })
    _save_campaigns(campaigns)

    return {"status": "sent", "sent_to": TEST_RECIPIENT}


# ── Lead Generation — Oregon Metro RLIS ──────────────────────────────────────

_GEOCODING_URL      = "https://maps.googleapis.com/maps/api/geocode/json"
_SOLAR_URL          = "https://solar.googleapis.com/v1/buildingInsights:findClosest"
_PLACES_FIND_URL    = "https://maps.googleapis.com/maps/api/place/findplacefromtext/json"
_PLACES_NEARBY_URL  = "https://maps.googleapis.com/maps/api/place/nearbysearch/json"
_PLACES_DETAILS_URL = "https://maps.googleapis.com/maps/api/place/details/json"
_PLACES_TEXT_URL    = "https://maps.googleapis.com/maps/api/place/textsearch/json"

# Search terms used to find commercial properties via Places Text Search
_COMMERCIAL_QUERIES = [
    "shopping center",
    "strip mall",
    "office building",
    "commercial property management",
]

# RLIS_DATA_DIR env var lets Render's persistent disk be used without changing DATA_DIR
_RLIS_DATA_DIR = Path(os.getenv("RLIS_DATA_DIR", str(DATA_DIR)))
_RLIS_DBF   = _RLIS_DATA_DIR / "TAXLOTS" / "taxlots_public.dbf"
_RLIS_CACHE = _RLIS_DATA_DIR / "rlis_cache.csv"

# Oregon North State Plane (NAD83 HARN, Intl Feet) → WGS84 (lng, lat)
_OR_PROJ = Transformer.from_crs("EPSG:2913", "EPSG:4326", always_xy=True)

_RESIDENTIAL_LU = {"SFR", "MFR"}
_COMMERCIAL_LU  = {"COM", "IND"}

_rlis_df: pd.DataFrame | None = None


def _read_rlis_dbf() -> pd.DataFrame:
    """Read taxlots_public.dbf and return a slim DataFrame with only the fields we need."""
    keep = {"SITEADDR", "SITECITY", "SITEZIP", "COUNTY", "LANDUSE", "YEARBUILT", "X_COORD", "Y_COORD"}
    records = []
    with open(_RLIS_DBF, "rb") as f:
        hdr = f.read(32)
        num_records  = struct.unpack("<I", hdr[4:8])[0]
        header_size  = struct.unpack("<H", hdr[8:10])[0]
        record_size  = struct.unpack("<H", hdr[10:12])[0]

        fields: list[tuple[str, int]] = []
        while True:
            fd = f.read(32)
            if fd[0] == 13:
                break
            name   = fd[:11].replace(b"\x00", b"").decode("ascii", errors="ignore").strip()
            length = fd[16]
            fields.append((name, length))

        f.seek(header_size)
        for _ in range(num_records):
            raw = f.read(record_size)
            if not raw or raw[0] == 42:   # deleted record
                continue
            row: dict[str, str] = {}
            pos = 1
            for name, length in fields:
                chunk = raw[pos : pos + length]
                if name in keep:
                    row[name] = chunk.decode("latin-1", errors="replace").strip()
                pos += length
            records.append(row)

    return pd.DataFrame(records)


def _get_rlis_df() -> pd.DataFrame:
    global _rlis_df
    if _rlis_df is not None:
        return _rlis_df

    if _RLIS_CACHE.exists():
        _rlis_df = pd.read_csv(_RLIS_CACHE, dtype=str).fillna("")
    else:
        if not _RLIS_DBF.exists():
            raise FileNotFoundError(
                "RLIS data not found. Unzip rlis_free_addresses.zip inside backend/data/ "
                "so that backend/data/TAXLOTS/taxlots_public.dbf exists."
            )
        _rlis_df = _read_rlis_dbf()
        _rlis_df.to_csv(_RLIS_CACHE, index=False)   # cache for fast future loads

    return _rlis_df


def _query_rlis(zip_code: str, county: str, prop_type: str, limit: int = 100) -> list[dict]:
    df = _get_rlis_df()
    base_mask = pd.Series(True, index=df.index)

    if zip_code:
        base_mask &= df["SITEZIP"].str[:5] == zip_code.strip()

    if county:
        county_map = {"multnomah": "M", "clackamas": "C", "washington": "W"}
        code = county_map.get(county.lower().strip())
        if not code:
            code = county.strip().upper()[:1]
        base_mask &= df["COUNTY"] == code

    if prop_type == "residential":
        filtered = df[base_mask & df["LANDUSE"].isin(_RESIDENTIAL_LU)].head(limit)
    elif prop_type == "commercial":
        filtered = df[base_mask & df["LANDUSE"].isin(_COMMERCIAL_LU)].head(limit)
    else:
        # "both" — split the budget evenly so commercial can't be crowded out by residential
        half = limit // 2
        res_rows = df[base_mask & df["LANDUSE"].isin(_RESIDENTIAL_LU)].head(half)
        com_rows = df[base_mask & df["LANDUSE"].isin(_COMMERCIAL_LU)].head(half)
        # If one side has fewer than half, let the other fill the remaining slots
        res_take = len(res_rows)
        com_take = len(com_rows)
        if res_take < half:
            com_rows = df[base_mask & df["LANDUSE"].isin(_COMMERCIAL_LU)].head(limit - res_take)
        elif com_take < half:
            res_rows = df[base_mask & df["LANDUSE"].isin(_RESIDENTIAL_LU)].head(limit - com_take)
        filtered = pd.concat([res_rows, com_rows])

    results = []
    for _, row in filtered.iterrows():
        try:
            x, y = float(row["X_COORD"]), float(row["Y_COORD"])
            if x == 0 or y == 0:
                continue
            lng, lat = _OR_PROJ.transform(x, y)
        except (ValueError, TypeError):
            continue

        lu = row.get("LANDUSE", "").strip().upper()
        ptype = "Residential" if lu in _RESIDENTIAL_LU else "Commercial" if lu in _COMMERCIAL_LU else "Other"

        try:
            year_built = int(float(row.get("YEARBUILT") or 0))
        except (ValueError, TypeError):
            year_built = 0

        addr  = row.get("SITEADDR", "").strip()
        city  = row.get("SITECITY", "").strip().title()
        szip  = row.get("SITEZIP",  "").strip()[:5]
        formatted = f"{addr}, {city}, OR {szip}".strip(", ")

        results.append({
            "address":    formatted,
            "lat":        lat,
            "lng":        lng,
            "property_type": ptype,
            "year_built": year_built,
            "lu_class":   lu,
        })

    return results


def _score_from_year_built(year_built: int) -> tuple[str, str]:
    if year_built <= 0:
        return "Low", "Year built unknown."
    age = date.today().year - year_built
    if age >= 15:
        return "High",   f"Built {year_built} — {age} yrs old, likely due for inspection."
    elif age >= 8:
        return "Medium", f"Built {year_built} — {age} yrs old, worth a follow-up."
    else:
        return "Low",    f"Built {year_built} — {age} yrs old, lower immediate priority."


def _score_from_solar_data(
    solar_data: dict,
    label: str = "",
    year_built_rlis: int | None = None,
) -> tuple[str, str]:
    """
    Derive roof priority from Solar API data.

    Priority order (highest confidence first):
      1. Solar API yearBuilt  — actual construction year if Google provides it
      2. RLIS yearBuilt       — county tax-record construction year
      3. Solar imageryDate    — satellite photo age (proxy; not the roof age)

    Logs which source was used for every property.
    """
    today = date.today()
    tag   = f"[score{' ' + label if label else ''}]"

    def _rate(yb: int, source: str) -> tuple[str, str]:
        age = today.year - yb
        print(f"{tag} source={source}, yearBuilt={yb}, age={age} yrs")
        if age >= 15:
            return "High",   f"Built {yb} — {age} yrs old, likely due for inspection."
        elif age >= 8:
            return "Medium", f"Built {yb} — {age} yrs old, worth a follow-up."
        else:
            return "Low",    f"Built {yb} — {age} yrs old, lower immediate priority."

    # 1. Solar API yearBuilt (not present in current API; included for future-proofing)
    raw_yb = (
        solar_data.get("yearBuilt")
        or (solar_data.get("solarPotential") or {}).get("yearBuilt")
    )
    if raw_yb:
        try:
            return _rate(int(raw_yb), "Solar.yearBuilt")
        except (ValueError, TypeError):
            pass

    # 2. RLIS county tax-record year_built
    if year_built_rlis and year_built_rlis > 1800:
        return _rate(year_built_rlis, "RLIS.yearBuilt")

    # 3. imageryDate fallback (satellite photo age, not roof age)
    imagery_date = solar_data.get("imageryDate")
    if imagery_date:
        yr  = imagery_date.get("year",  today.year)
        mo  = imagery_date.get("month", 1)
        age = ((today.year - yr) * 12 + (today.month - mo)) // 12
        print(f"{tag} source=imageryDate, date={yr}-{mo:02d}, imagery age={age} yrs (note: NOT building age)")
        if age >= 15:
            return "High",   f"Imagery from {yr} — roof likely overdue for inspection."
        elif age >= 8:
            return "Medium", f"Imagery from {yr} — worth scheduling a follow-up."
        else:
            return "Low",    f"Recent imagery from {yr} — lower immediate priority."

    print(f"{tag} no scoring data available")
    return "Unknown", "Building age could not be determined."


def _lookup_rlis_year_built(address: str) -> int | None:
    """
    Look up YEARBUILT in the RLIS cache by matching street number + zip code.
    Returns None if RLIS is unavailable or no match is found.
    """
    try:
        zip_m = re.search(r'\b(\d{5})\b', address)
        num_m = re.match(r'^\s*(\d+)\s', address.strip())
        if not zip_m or not num_m:
            return None
        zip_code   = zip_m.group(1)
        street_num = num_m.group(1)

        df = _get_rlis_df()
        candidates = df[
            (df["SITEZIP"].str[:5] == zip_code)
            & df["SITEADDR"].str.startswith(street_num + " ")
        ]
        if candidates.empty:
            return None

        raw = str(candidates.iloc[0].get("YEARBUILT", "") or "").strip()
        yb  = int(float(raw)) if raw else 0
        return yb if yb > 1800 else None
    except Exception:
        return None


@app.get("/roof-score")
def roof_score(address: str):
    api_key = os.getenv("GOOGLE_SOLAR_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="GOOGLE_SOLAR_API_KEY not configured.")

    try:
        r = httpx.get(_GEOCODING_URL, params={"address": address, "key": api_key}, timeout=10)
        r.raise_for_status()
        geo = r.json()
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Geocoding error: {e}")

    if geo.get("status") != "OK" or not geo.get("results"):
        status = geo.get("status", "UNKNOWN")
        msg = geo.get("error_message", "")
        raise HTTPException(status_code=404, detail=f"Geocoding failed: {status}. {msg}".strip())

    loc = geo["results"][0]["geometry"]["location"]
    lat, lng = loc["lat"], loc["lng"]

    try:
        r = httpx.get(_SOLAR_URL, params={
            "location.latitude": lat,
            "location.longitude": lng,
            "key": api_key,
            "requiredQuality": "LOW",
        }, timeout=12)
        if r.status_code == 404:
            raise HTTPException(status_code=404, detail="No solar data found for this address.")
        r.raise_for_status()
        solar = r.json()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Solar API error: {e}")

    imagery_date = solar.get("imageryDate")
    if not imagery_date:
        raise HTTPException(status_code=404, detail="No imagery data available for this address.")

    year  = imagery_date.get("year", date.today().year)
    month = imagery_date.get("month", 1)
    today = date.today()
    age_years = ((today.year - year) * 12 + (today.month - month)) // 12
    if age_years >= 15:
        score, summary = "High",   f"Imagery from {year} — roof likely overdue for inspection."
    elif age_years >= 8:
        score, summary = "Medium", f"Imagery from {year} — worth scheduling a follow-up."
    else:
        score, summary = "Low",    f"Recent imagery from {year} — lower immediate priority."

    return {
        "address":      geo["results"][0]["formatted_address"],
        "score":        score,
        "age_years":    age_years,
        "imagery_date": f"{year}-{month:02d}",
        "summary":      summary,
    }


class GenerateLeadsRequest(BaseModel):
    zip:           str = ""
    county:        str = ""
    property_type: str = "both"


def _text_search_places(query_prefix: str, location: str, google_key: str) -> list[dict]:
    """Run one Google Places Text Search and return basic place info (no phone/website yet)."""
    try:
        r = httpx.get(
            _PLACES_TEXT_URL,
            params={"query": f"{query_prefix} {location}", "region": "us", "key": google_key},
            timeout=10,
        )
        results = (r.json() if r.status_code == 200 else {}).get("results", [])
        places = []
        for p in results:
            pid = p.get("place_id", "")
            if not pid:
                continue
            loc = p.get("geometry", {}).get("location", {})
            places.append({
                "place_id": pid,
                "name":     p.get("name", ""),
                "address":  p.get("formatted_address", ""),
                "lat":      loc.get("lat", 0.0),
                "lng":      loc.get("lng", 0.0),
            })
        return places
    except Exception as e:
        print(f"[_text_search_places] Error for '{query_prefix} {location}': {e}")
        return []


def _validate_website(url: str) -> str | None:
    """Return url if it looks like a real HTTP URL, otherwise None."""
    if url and url.startswith(("http://", "https://")):
        return url
    return None


def _enrich_place(place: dict, google_key: str, apollo_key: str) -> dict | None:
    """
    Step 2: Place Details — get phone + website for the confirmed business.
    Step 3: Solar API — score the roof from satellite imagery age.
    Step 4: Apollo — decision-maker email via org domain or business name.
    Never raises; returns None only on a total Place Details failure.
    """
    place_id      = place["place_id"]
    business_name = place["name"]
    address       = place["address"]
    lat           = place["lat"]
    lng           = place["lng"]
    phone = ""
    website = ""

    # Place Details
    try:
        rd = httpx.get(
            _PLACES_DETAILS_URL,
            params={"place_id": place_id,
                    "fields": "name,formatted_address,formatted_phone_number,website,geometry",
                    "key": google_key},
            timeout=8,
        )
        detail = (rd.json() if rd.status_code == 200 else {}).get("result", {})
        if detail.get("name"):
            business_name = detail["name"]
        if detail.get("formatted_address"):
            address = detail["formatted_address"]
        phone   = detail.get("formatted_phone_number", "")
        raw_site = detail.get("website", "")
        print(f"[website] raw from Places: '{raw_site}' for {business_name}")
        website = _validate_website(raw_site)
        loc = detail.get("geometry", {}).get("location", {})
        if loc:
            lat = loc.get("lat", lat)
            lng = loc.get("lng", lng)
    except Exception as e:
        print(f"[_enrich_place] Place Details error for {place_id}: {e}")

    # RLIS lookup — county tax records often have the actual construction year
    year_built = _lookup_rlis_year_built(address)
    if year_built:
        print(f"[_enrich_place] {business_name}: RLIS year_built={year_built}")

    # Solar API — prefer yearBuilt > RLIS > imageryDate
    score   = "Unknown"
    summary = ""
    try:
        r = httpx.get(_SOLAR_URL, params={
            "location.latitude":  lat,
            "location.longitude": lng,
            "key": google_key,
            "requiredQuality": "LOW",
        }, timeout=12)
        if r.status_code == 200:
            score, summary = _score_from_solar_data(
                r.json(), label=business_name, year_built_rlis=year_built
            )
        elif year_built:
            # Solar unavailable but we have RLIS — still score it
            score, summary = _score_from_year_built(year_built)
            print(f"[_enrich_place] {business_name}: Solar 404, using RLIS year_built={year_built}")
    except Exception as e:
        print(f"[_enrich_place] Solar API error for {business_name}: {e}")
        if year_built:
            score, summary = _score_from_year_built(year_built)

    # Apollo — org enrich for phone fallback + people search for decision-maker email
    email = ""
    contact_name = ""
    if apollo_key and (website or business_name):
        try:
            if website:
                domain = website.split("//")[-1].split("/")[0]
                if domain.startswith("www."):
                    domain = domain[4:]
                org_payload: dict = {"domain": domain}
            else:
                org_payload = {"name": business_name}
            r = httpx.post(
                "https://api.apollo.io/api/v1/organizations/enrich",
                headers={"X-Api-Key": apollo_key, "Content-Type": "application/json"},
                json=org_payload, timeout=8,
            )
            if r.status_code == 200:
                org = r.json().get("organization") or {}
                if not phone:
                    p = org.get("phone") or (org.get("primary_phone") or {}).get("number", "")
                    if p:
                        phone = p
        except Exception:
            pass

        if business_name and not email:
            person: dict = {}
            try:
                r = httpx.post(
                    "https://api.apollo.io/api/v1/people/search",
                    headers={"X-Api-Key": apollo_key, "Content-Type": "application/json"},
                    json={"organization_name": business_name,
                          "titles": ["property manager", "owner", "facilities manager"],
                          "page": 1, "per_page": 1},
                    timeout=8,
                )
                if r.status_code == 200:
                    people = r.json().get("people", [])
                    person = people[0] if people else {}
                else:
                    r2 = httpx.post(
                        "https://api.apollo.io/api/v1/people/match",
                        headers={"X-Api-Key": apollo_key, "Content-Type": "application/json"},
                        json={"organization_name": business_name}, timeout=8,
                    )
                    if r2.status_code == 200:
                        person = r2.json().get("person") or {}
            except Exception:
                pass
            if person:
                e = person.get("email", "")
                if e and "@" in e:
                    email = e
                first = person.get("first_name", "")
                last  = person.get("last_name", "")
                n = f"{first} {last}".strip()
                if n:
                    contact_name = n

    return {
        "address":       address,
        "business_name": business_name,
        "phone":         phone,
        "email":         email,
        "contact_name":  contact_name,
        "website":       website or None,
        "score":         score,
        "summary":       summary,
        "year_built":    year_built,
        "property_type": "Commercial",
        "enriched":      True,
    }


@app.post("/generate-leads")
def generate_leads(req: GenerateLeadsRequest):
    if not req.zip.strip() and not req.county.strip():
        raise HTTPException(status_code=400, detail="Zip code or county required.")

    google_key = os.getenv("GOOGLE_SOLAR_API_KEY", "")
    apollo_key = os.getenv("APOLLO_API_KEY", "")
    score_order = {"High": 0, "Medium": 1, "Low": 2, "Unknown": 3}
    leads: list[dict] = []

    # ── Commercial: Google Places Text Search pipeline ─────────────────────────
    # Start from real active businesses, not raw property parcels.
    if req.property_type in ("commercial", "both"):
        if not google_key:
            raise HTTPException(status_code=500, detail="GOOGLE_SOLAR_API_KEY not configured.")

        # Build location label for the search queries
        location = req.zip.strip() or f"{req.county.strip().title()} County Oregon"

        # Step 1 — Run all 4 text search queries concurrently, deduplicate by place_id
        with ThreadPoolExecutor(max_workers=4) as executor:
            search_futures = {
                executor.submit(_text_search_places, q, location, google_key): q
                for q in _COMMERCIAL_QUERIES
            }
            seen_ids: set[str] = set()
            places: list[dict] = []
            for f in list(search_futures):
                try:
                    for p in f.result(timeout=15):
                        if p["place_id"] not in seen_ids:
                            seen_ids.add(p["place_id"])
                            places.append(p)
                except Exception as e:
                    print(f"[generate_leads] Text search error: {e}")

        # Steps 2–4 — Place Details + Solar + Apollo for each place, all concurrent
        if places:
            with ThreadPoolExecutor(max_workers=20) as executor:
                enrich_futures = {
                    executor.submit(_enrich_place, p, google_key, apollo_key): p["place_id"]
                    for p in places
                }
                for f in list(enrich_futures):
                    try:
                        result = f.result(timeout=25)
                        if result:
                            leads.append(result)
                    except Exception as e:
                        print(f"[generate_leads] Enrich error: {e}")

    # ── Residential: RLIS pipeline (unchanged) ─────────────────────────────────
    if req.property_type in ("residential", "both"):
        try:
            addresses = _query_rlis(req.zip.strip(), req.county.strip(), "residential", limit=250)
            for a in addresses:
                score, summary = _score_from_year_built(a["year_built"])
                leads.append({
                    "address":       a["address"],
                    "property_type": "Residential",
                    "year_built":    a["year_built"] if a["year_built"] else None,
                    "score":         score,
                    "summary":       summary,
                    "enriched":      False,
                    "business_name": "",
                    "phone":         "",
                    "email":         "",
                    "contact_name":  "",
                    "website":       None,
                })
        except FileNotFoundError as e:
            if req.property_type == "residential":
                raise HTTPException(status_code=500, detail=str(e))
            # For "both", continue with whatever commercial leads we found

    if not leads:
        raise HTTPException(
            status_code=404,
            detail="No results found. Try a different zip code or county.",
        )

    leads.sort(key=lambda x: score_order.get(x["score"], 4))
    if leads:
        print(f"[generate-leads] first lead JSON: {json.dumps(leads[0], default=str)}")
    return {"leads": leads}


class ScoreAddressesRequest(BaseModel):
    addresses: list[str]


@app.post("/score-addresses")
def score_addresses(req: ScoreAddressesRequest):
    api_key = os.getenv("GOOGLE_SOLAR_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="GOOGLE_SOLAR_API_KEY not configured.")

    lines = [a.strip() for a in req.addresses if a.strip()]
    if not lines:
        raise HTTPException(status_code=400, detail="No addresses provided.")

    score_order = {"High": 0, "Medium": 1, "Low": 2, "Unknown": 3}
    results = []

    for address in lines:
        try:
            r = httpx.get(_GEOCODING_URL, params={"address": address, "key": api_key}, timeout=10)
            r.raise_for_status()
            geo = r.json()
        except Exception:
            results.append({"address": address, "score": "Unknown", "summary": "Geocoding request failed.", "age_years": None, "imagery_date": None})
            continue

        if geo.get("status") != "OK" or not geo.get("results"):
            results.append({"address": address, "score": "Unknown", "summary": "Address not found.", "age_years": None, "imagery_date": None})
            continue

        loc = geo["results"][0]["geometry"]["location"]
        formatted = geo["results"][0]["formatted_address"]

        try:
            r = httpx.get(_SOLAR_URL, params={
                "location.latitude": loc["lat"],
                "location.longitude": loc["lng"],
                "key": api_key,
                "requiredQuality": "LOW",
            }, timeout=12)
            if r.status_code == 404:
                results.append({"address": formatted, "score": "Unknown", "summary": "No solar data available.", "age_years": None, "imagery_date": None})
                continue
            r.raise_for_status()
            solar = r.json()
        except HTTPException:
            raise
        except Exception:
            results.append({"address": formatted, "score": "Unknown", "summary": "Solar API request failed.", "age_years": None, "imagery_date": None})
            continue

        imagery_date = solar.get("imageryDate")
        if not imagery_date:
            results.append({"address": formatted, "score": "Unknown", "summary": "No imagery data available.", "age_years": None, "imagery_date": None})
            continue

        year = imagery_date.get("year", date.today().year)
        month = imagery_date.get("month", 1)
        today = date.today()
        age_years = ((today.year - year) * 12 + (today.month - month)) // 12

        if age_years >= 15:
            score, summary = "High",   f"Imagery from {year} — roof likely overdue for inspection."
        elif age_years >= 8:
            score, summary = "Medium", f"Imagery from {year} — worth scheduling a follow-up."
        else:
            score, summary = "Low",    f"Recent imagery from {year} — lower immediate priority."

        results.append({"address": formatted, "score": score, "summary": summary, "age_years": age_years, "imagery_date": f"{year}-{month:02d}"})

    results.sort(key=lambda x: score_order.get(x["score"], 3))
    return {"results": results}


_RESIDENTIAL_TYPES  = {"residential", "sfr", "mfr"}
_STREET_SUFFIX_RE = re.compile(
    r'\b(ave|avenue|st|street|rd|road|blvd|boulevard|dr|drive|'
    r'ln|lane|ct|court|pl|place|way|pkwy|parkway|hwy|highway)\b',
    re.IGNORECASE,
)


def _is_address_like(name: str) -> bool:
    """True if Places returned an address/ZIP instead of a real business name."""
    name = name.strip()
    if re.match(r'^\d{5}$', name):                          # bare ZIP
        return True
    if re.match(r'^\d+\s', name) and _STREET_SUFFIX_RE.search(name):   # "123 Main St"
        return True
    if re.match(r'^\d+[-–]\d+\s', name) and _STREET_SUFFIX_RE.search(name):  # "3520-3524 N Mississippi Ave"
        return True
    return False


class EnrichRequest(BaseModel):
    name: str = ""
    company: str = ""
    address: str = ""
    phone: str = ""
    email: str = ""
    property_type: str = ""


def _parse_address(address: str) -> dict:
    """Split 'street, city, ST zip' into components for BatchData."""
    parts = [p.strip() for p in address.split(",")]
    street = parts[0] if len(parts) > 0 else address
    city   = parts[1] if len(parts) > 1 else ""
    state  = ""
    zip_   = ""
    if len(parts) > 2:
        # last part could be "OR 97201" or just "OR"
        tail = parts[-1].strip().split()
        if len(tail) >= 2:
            state, zip_ = tail[0], tail[1]
        elif tail:
            state = tail[0]
    return {"street": street, "city": city, "state": state, "zip": zip_}


@app.post("/enrich")
def enrich_contact(req: EnrichRequest):
    phone = req.phone.strip()
    email = req.email.strip()
    name  = req.name.strip()
    company = req.company.strip()
    phone_found = False
    email_found = False

    # ── Step 1: BatchData Skip Trace — owner name, phone, email ──────────────
    # Works for both residential and commercial; covers the gap where
    # Google Places and Apollo can't resolve anonymous RLIS addresses.
    batchdata_key = os.getenv("BATCHDATA_API_KEY", "")
    if batchdata_key and batchdata_key != "your_batchdata_api_key_here":
        try:
            addr = _parse_address(req.address)
            r = httpx.post(
                "https://api.batchdata.com/api/v1/property/skip-trace",
                headers={
                    "Authorization": f"Bearer {batchdata_key}",
                    "Content-Type": "application/json",
                },
                json={"requests": [{"propertyAddress": addr}]},
                timeout=15,
            )
            if r.status_code == 200:
                results = r.json().get("results", {}).get("properties", [])
                if results:
                    prop = results[0]
                    owners = prop.get("owners") or []
                    owner  = owners[0] if owners else {}

                    if not name:
                        first = owner.get("firstName", "")
                        last  = owner.get("lastName", "")
                        full  = f"{first} {last}".strip()
                        if full:
                            name = full
                            print(f"[enrich] BatchData → name: {name}")

                    if not phone:
                        for ph in owner.get("phones", []):
                            num = ph.get("number", "") or ph.get("phone", "")
                            if num:
                                phone = num
                                phone_found = True
                                print(f"[enrich] BatchData → phone: {phone}")
                                break

                    if not email:
                        for em in owner.get("emails", []):
                            addr_val = em.get("email", "") or em.get("address", "")
                            if addr_val and "@" in addr_val:
                                email = addr_val
                                email_found = True
                                print(f"[enrich] BatchData → email: {email}")
                                break
                else:
                    print("[enrich] BatchData → no results")
            else:
                print(f"[enrich] BatchData → HTTP {r.status_code}: {r.text[:200]}")
        except Exception as e:
            print(f"[enrich] BatchData error: {e}")
    else:
        print("[enrich] BatchData key not configured — skipping")

    # Residential: BatchData is sufficient, skip Google Places / Apollo
    if req.property_type.strip().lower() in _RESIDENTIAL_TYPES:
        print(f"[enrich] Residential done — phone_found={phone_found}, email_found={email_found}")
        return {
            "name": name, "company": company, "address": req.address,
            "phone": phone, "email": email,
            "email_found": email_found, "phone_found": phone_found,
            "enrichment_run": True,
        }

    # ── Step 2: Google Places — business name + phone for commercial ──────────
    google_key = os.getenv("GOOGLE_SOLAR_API_KEY", "")
    website    = ""
    if google_key and (not phone or not company):
        try:
            query = f"{company} {req.address}".strip() if company else req.address
            r = httpx.get(
                _PLACES_FIND_URL,
                params={"input": query, "inputtype": "textquery",
                        "fields": "name,place_id", "key": google_key},
                timeout=10,
            )
            body = r.json() if r.status_code == 200 else {}
            candidates = body.get("candidates", [])
            if candidates:
                hit      = candidates[0]
                place_id = hit.get("place_id", "")
                if not company:
                    n = hit.get("name", "")
                    if n and not _is_address_like(n):
                        company = n
                if place_id:
                    rd = httpx.get(
                        _PLACES_DETAILS_URL,
                        params={"place_id": place_id,
                                "fields": "name,formatted_phone_number,website",
                                "key": google_key},
                        timeout=10,
                    )
                    detail = (rd.json() if rd.status_code == 200 else {}).get("result", {})
                    if not company:
                        n = detail.get("name", "")
                        if n and not _is_address_like(n):
                            company = n
                    if not phone:
                        p = detail.get("formatted_phone_number", "")
                        if p:
                            phone = p
                            phone_found = True
                            print(f"[enrich] Places → phone: {phone}")
                    website = detail.get("website", "")
                    print(f"[enrich] Places → company: {company or 'none'}, website: {website or 'none'}")
            else:
                print(f"[enrich] Places → {body.get('status', 'no candidates')}")
        except Exception as e:
            print(f"[enrich] Places error: {e}")

    # ── Step 3: Apollo — fallback email via org domain ────────────────────────
    apollo_key = os.getenv("APOLLO_API_KEY", "")
    if apollo_key and not email and (website or company):
        try:
            if website:
                domain = website.split("//")[-1].split("/")[0]
                if domain.startswith("www."):
                    domain = domain[4:]
                org_payload: dict = {"domain": domain}
            else:
                org_payload = {"name": company}
            r = httpx.post(
                "https://api.apollo.io/api/v1/organizations/enrich",
                headers={"X-Api-Key": apollo_key, "Content-Type": "application/json"},
                json=org_payload, timeout=10,
            )
            if r.status_code == 200:
                org = r.json().get("organization") or {}
                if not phone:
                    p = org.get("phone") or (org.get("primary_phone") or {}).get("number", "")
                    if p:
                        phone = p
                        phone_found = True
                        print(f"[enrich] Apollo org → phone: {phone}")
        except Exception as e:
            print(f"[enrich] Apollo org error: {e}")

    print(f"[enrich] Done — phone_found={phone_found}, email_found={email_found}")
    return {
        "name": name,
        "company": company,
        "address": req.address,
        "phone": phone,
        "email": email,
        "email_found": email_found,
        "phone_found": phone_found,
        "enrichment_run": True,
    }


class ContactIn(BaseModel):
    name: str = ""
    company: str = ""
    phone: str = ""
    email: str = ""
    address: str = ""
    property_type: str = ""
    priority_tier: str = ""


@app.post("/contacts/add")
def add_contact(contact: ContactIn):
    existing = _load_contacts()
    existing_emails = {c.get("email", "").strip().lower() for c in existing if c.get("email")}

    email = contact.email.strip().lower()
    if email and email in existing_emails:
        raise HTTPException(status_code=409, detail="A contact with this email already exists.")

    row = {**contact.model_dump(), "last_contacted": None}
    existing.append(row)
    _save_contacts(existing)
    return {"status": "added"}


# ── Time Tracking ──────────────────────────────────────────────────────────────

_SHEET_ID   = "1MzNRiBW2-RIaCJ-BIqb7f_Faj3Ej2P1S4e4XgOmoqVY"
_CREDS_FILE = Path(__file__).parent / "google-credentials.json"
_SCOPES     = ["https://www.googleapis.com/auth/spreadsheets"]

_gs_client = None
_gs_sheet  = None


def _get_sheet():
    global _gs_client, _gs_sheet
    if _gs_sheet is None:
        creds_json = os.getenv("GOOGLE_CREDENTIALS_JSON")
        if creds_json:
            import json as _json
            info   = _json.loads(creds_json)
            creds  = _SACredentials.from_service_account_info(info, scopes=_SCOPES)
        else:
            creds  = _SACredentials.from_service_account_file(str(_CREDS_FILE), scopes=_SCOPES)
        _gs_client = gspread.authorize(creds)
        _gs_sheet  = _gs_client.open_by_key(_SHEET_ID)
    return _gs_sheet


def _employees_ws():
    return _get_sheet().worksheet("Employees")


def _timesheets_ws():
    return _get_sheet().worksheet("Timesheets")


def _norm(records: list[dict]) -> list[dict]:
    """Normalise Sheets records: lowercase+underscore keys, empty strings → None."""
    out = []
    for row in records:
        out.append({
            k.strip().lower().replace(" ", "_"): (None if v == "" else v)
            for k, v in row.items()
        })
    return out


def _load_employees() -> list[dict]:
    try:
        return _norm(_employees_ws().get_all_records())
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"Could not read Employees sheet: {exc}")


def _append_employee(emp: dict) -> None:
    try:
        _employees_ws().append_row(
            [emp["id"], emp["name"], emp["token"]],
            value_input_option="RAW",
        )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"Could not write to Employees sheet: {exc}")


def _load_timesheets() -> list[dict]:
    try:
        return _norm(_timesheets_ws().get_all_records())
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"Could not read Timesheets sheet: {exc}")


def _append_timesheet(entry: dict) -> None:
    # Columns: id | employee_id | employee_name | clock_in | lunch_out | lunch_in | clock_out | lunch_hours | work_hours | notes
    try:
        _timesheets_ws().append_row(
            [entry["id"], entry["employee_id"], entry["employee_name"],
             entry["clock_in"], "", "", "", "", "", ""],
            value_input_option="RAW",
        )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"Could not write to Timesheets sheet: {exc}")


def _update_cell(entry_id: str, col: int, value) -> None:
    """Update a single cell in the Timesheets sheet, located by id in column 1."""
    try:
        ws   = _timesheets_ws()
        cell = ws.find(entry_id, in_column=1)
        if cell:
            ws.update_cell(cell.row, col, value)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"Could not update Timesheets sheet: {exc}")


def _compute_hours(clock_in_full: str, clock_out_time: str,
                   lunch_out: str = "", lunch_in: str = "") -> tuple[float, float]:
    """Return (lunch_hours, work_hours) as decimal hours using HH:MM precision (no seconds)."""
    date_str  = clock_in_full[:10]
    ci = datetime.strptime(clock_in_full.strip(), "%Y-%m-%d %H:%M")
    co = datetime.strptime(f"{date_str} {clock_out_time.strip()}", "%Y-%m-%d %H:%M")
    if co < ci:
        co += timedelta(days=1)
    gross_min = int((co - ci).total_seconds() // 60)

    lunch_min = 0
    if lunch_out.strip() and lunch_in.strip():
        lo = datetime.strptime(f"{date_str} {lunch_out.strip()}", "%Y-%m-%d %H:%M")
        li = datetime.strptime(f"{date_str} {lunch_in.strip()}", "%Y-%m-%d %H:%M")
        if li < lo:
            li += timedelta(days=1)
        lunch_min = int((li - lo).total_seconds() // 60)

    return round(lunch_min / 60, 2), round((gross_min - lunch_min) / 60, 2)


def _update_clock_out(entry_id: str, clock_out: str, clock_in_full: str,
                      lunch_out: str = "", lunch_in: str = "", notes: str = "") -> None:
    """Write clock_out (col 7), lunch_hours (col 8), work_hours (col 9), and optionally notes (col 10)."""
    lunch_hours, work_hours = 0.0, 0.0
    try:
        lunch_hours, work_hours = _compute_hours(clock_in_full, clock_out, lunch_out, lunch_in)
    except Exception:
        pass
    try:
        ws   = _timesheets_ws()
        cell = ws.find(entry_id, in_column=1)
        if cell:
            ws.update_cell(cell.row, 7, clock_out)    # col 7 = clock_out
            ws.update_cell(cell.row, 8, lunch_hours)  # col 8 = lunch_hours
            ws.update_cell(cell.row, 9, work_hours)   # col 9 = work_hours
            if notes:
                ws.update_cell(cell.row, 10, notes)   # col 10 = notes
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"Could not update Timesheets sheet: {exc}")


def _make_qr_b64(url: str) -> str:
    qr = _qrlib.QRCode(
        version=1,
        error_correction=_qrlib.constants.ERROR_CORRECT_M,
        box_size=10,
        border=4,
    )
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return base64.b64encode(buf.getvalue()).decode()


def _row_date(t: dict) -> str:
    """Return the date string for a timesheet row, falling back to parsing clock_in."""
    d = t.get("date")
    if d:
        return str(d)
    ci = t.get("clock_in") or ""
    return ci[:10]  # ISO timestamp → "YYYY-MM-DD"


def _today_status(employee_id: str, timesheets: list[dict]) -> str:
    """Return one of: not_clocked_in | clocked_in | on_lunch | returned_from_lunch.

    Uses the most recent row for today. A completed row (clock_out filled) means
    the employee can start a new session, so we return not_clocked_in.
    Multiple rows per employee per day are valid.
    """
    today = datetime.now(_TZ).date().isoformat()
    today_rows = [
        t for t in timesheets
        if t.get("employee_id") == employee_id and _row_date(t) == today
    ]
    if not today_rows:
        return "not_clocked_in"
    row = today_rows[-1]  # most recent session
    if not _is_blank(row.get("clock_out")):
        return "not_clocked_in"  # last session complete — ready for a new one
    if not _is_blank(row.get("lunch_in")):
        return "returned_from_lunch"
    if not _is_blank(row.get("lunch_out")):
        return "on_lunch"
    return "clocked_in"


def _fmt_clock(time_str: str) -> str:
    """Convert 'HH:MM', 'YYYY-MM-DD HH:MM', or ISO timestamp to '9:30 AM' display format."""
    try:
        s = time_str.strip()
        if "T" in s or (len(s) > 5 and " " in s):
            dt = datetime.fromisoformat(s.replace(" ", "T"))
            h, m = dt.hour, dt.minute
        else:
            h, m = (int(p) for p in s.split(":")[:2])
        ampm = "AM" if h < 12 else "PM"
        h12  = h % 12 or 12
        return f"{h12}:{m:02d} {ampm}"
    except Exception:
        return time_str


def _calc_hours(e: dict) -> float | None:
    """Calculate gross hours worked (clock_out - clock_in). clock_in='YYYY-MM-DD HH:MM', clock_out='HH:MM'."""
    ci_str = (e.get("clock_in") or "").strip()
    co_str = (e.get("clock_out") or "").strip()
    if not (ci_str and co_str):
        return None
    try:
        ci = datetime.strptime(ci_str, "%Y-%m-%d %H:%M")
        co = datetime.strptime(f"{ci_str[:10]} {co_str}", "%Y-%m-%d %H:%M")
        if co < ci:
            co += timedelta(days=1)
        return round((co - ci).total_seconds() / 3600, 2)
    except Exception:
        return None


def _calc_lunch(e: dict) -> int | None:
    """Calculate lunch duration in minutes from lunch_out and lunch_in (both 'HH:MM')."""
    lo_str = (e.get("lunch_out") or "").strip()
    li_str = (e.get("lunch_in")  or "").strip()
    if not (lo_str and li_str):
        return None
    try:
        date_str = _row_date(e)
        lo = datetime.strptime(f"{date_str} {lo_str}", "%Y-%m-%d %H:%M")
        li = datetime.strptime(f"{date_str} {li_str}", "%Y-%m-%d %H:%M")
        if li < lo:
            li += timedelta(days=1)
        return int((li - lo).total_seconds() / 60)
    except Exception:
        return None


def _clock_html(name: str, action: str, time_str: str,
                color: str = "#3b82f6", is_error: bool = False) -> str:
    icon     = "❌" if is_error else "⏰"
    err_attr = ' class="err"' if is_error else ""
    t_html   = f"<p class='time'>{time_str}</p>" if time_str else ""
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Time Clock</title>
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f8fafc;display:flex;align-items:center;justify-content:center;min-height:100vh;padding:24px}}
  .card{{background:#fff;border-radius:20px;padding:40px 32px;text-align:center;max-width:360px;width:100%;box-shadow:0 4px 24px rgba(0,0,0,.08)}}
  .co{{font-size:13px;color:#94a3b8;letter-spacing:.05em;text-transform:uppercase;margin-bottom:24px}}
  .icon{{font-size:56px;margin-bottom:16px}}
  .name{{font-size:26px;font-weight:700;color:#1e293b;margin-bottom:12px}}
  .action{{font-size:20px;font-weight:600;color:{color};margin-bottom:8px}}
  .time{{font-size:15px;color:#64748b}}
  .err{{color:#ef4444}}
</style>
</head>
<body>
<div class="card">
  <div class="co">Willamette Power Roofing</div>
  <div class="icon">{icon}</div>
  <div class="name"{err_attr}>{name}</div>
  <div class="action">{action}</div>
  {t_html}
</div>
</body>
</html>"""


class AddEmployeeRequest(BaseModel):
    name: str


@app.get("/employees")
def list_employees():
    employees  = _load_employees()
    timesheets = _load_timesheets()
    today      = datetime.now(_TZ).date().isoformat()

    # Midnight reset: auto-close any open entries from prior days
    stale = [
        t for t in timesheets
        if _row_date(t) < today and _is_blank(t.get("clock_out"))
    ]
    if stale:
        ts_ws = _timesheets_ws()
        for t in stale:
            cell = ts_ws.find(t.get("id", ""), in_column=1)
            if cell:
                co_time = "23:59"
                ci_full = (t.get("clock_in")  or "").strip()
                lo_str  = (t.get("lunch_out") or "").strip()
                li_str  = (t.get("lunch_in")  or "").strip()
                lunch_h, work_h = 0.0, 0.0
                try:
                    if ci_full:
                        lunch_h, work_h = _compute_hours(ci_full, co_time, lo_str, li_str)
                except Exception:
                    pass
                ts_ws.update_cell(cell.row, 7,  co_time)
                ts_ws.update_cell(cell.row, 8,  lunch_h)
                ts_ws.update_cell(cell.row, 9,  work_h)
                ts_ws.update_cell(cell.row, 10, "missed clock-out")
        timesheets = _load_timesheets()  # reload with closed entries

    result = []
    for emp in employees:
        status = _today_status(emp["id"], timesheets)
        today_rows = [
            t for t in timesheets
            if t.get("employee_id") == emp["id"] and _row_date(t) == today
        ]
        # Use the most recent open row for display times (None if no open session)
        open_row = next(
            (t for t in reversed(today_rows) if _is_blank(t.get("clock_out"))),
            None,
        )
        result.append({
            **emp,
            "status":         status,
            "clock_in_time":  open_row.get("clock_in")   if open_row else None,
            "lunch_out_time": open_row.get("lunch_out")  if open_row else None,
        })
    return {"employees": result}


@app.post("/employees")
def create_employee(req: AddEmployeeRequest):
    if not req.name.strip():
        raise HTTPException(status_code=400, detail="Name is required.")
    employees = _load_employees()
    if any(e["name"].lower() == req.name.strip().lower() for e in employees):
        raise HTTPException(status_code=400, detail=f"'{req.name.strip()}' already exists.")
    emp = {"id": str(uuid.uuid4()), "name": req.name.strip(), "token": secrets.token_urlsafe(8)}
    _append_employee(emp)
    return {**emp, "status": "Out"}


@app.get("/clock/{token}", response_class=HTMLResponse)
def clock_event(token: str):
    employees = _load_employees()
    emp = next((e for e in employees if e["token"] == token), None)
    if not emp:
        return HTMLResponse(
            _clock_html("Unknown", "Invalid QR code", "", is_error=True),
            status_code=404,
        )

    timesheets    = _load_timesheets()
    now_dt        = datetime.now(_TZ)
    today         = now_dt.date().isoformat()
    now_time      = now_dt.strftime("%H:%M")
    now_disp      = _fmt_clock(now_time)
    clock_in_full = f"{today} {now_time}"

    open_entry = next(
        (t for t in timesheets
         if t.get("employee_id") == emp["id"] and _is_blank(t.get("clock_out"))),
        None,
    )

    if open_entry:
        ci_full = (open_entry.get("clock_in") or "").strip()
        _update_clock_out(open_entry.get("id", ""), now_time, ci_full)
        return HTMLResponse(_clock_html(emp["name"], "Clocked Out", now_disp, color="#ef4444"))

    _append_timesheet({
        "id":            str(uuid.uuid4()),
        "employee_id":   emp["id"],
        "employee_name": emp["name"],
        "clock_in":      clock_in_full,
    })
    return HTMLResponse(_clock_html(emp["name"], "Clocked In", now_disp, color="#22c55e"))


class TimeclockActionRequest(BaseModel):
    employee_id: str
    action: str  # "clock_in" | "clock_out" | "lunch_out" | "lunch_in"


@app.get("/timeclock", response_class=HTMLResponse)
def timeclock_page():
    return HTMLResponse("""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Time Clock</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f8fafc;min-height:100vh;display:flex;flex-direction:column}
header{background:#1e3a5f;color:#fff;padding:16px 24px;text-align:center}
header h1{font-size:1.2rem;font-weight:700}
header p{font-size:.8rem;opacity:.7;margin-top:2px}
#session-bar{padding:9px 16px;text-align:center;font-size:.82rem;font-weight:500;background:#f0fdf4;border-bottom:1px solid #bbf7d0;color:#16a34a;min-height:36px}
#session-bar.expired{background:#fef2f2;border-color:#fecaca;color:#dc2626}
#error-banner{display:none;background:#fef2f2;border-bottom:2px solid #fca5a5;color:#dc2626;padding:14px 20px;text-align:center;font-size:1rem;font-weight:600;line-height:1.4}

.screen{display:none;flex-direction:column;flex:1;padding:28px 20px}
.screen.active{display:flex}

/* Screen 1 */
#s-who h2{text-align:center;font-size:1.3rem;font-weight:700;color:#1e293b;margin-bottom:6px}
#s-who .sub{text-align:center;font-size:1rem;color:#94a3b8;margin-bottom:28px;line-height:1.5}
.emp-btn{width:100%;background:#fff;border:1.5px solid #e2e8f0;border-radius:14px;padding:0 24px;margin-bottom:14px;font-size:1.2rem;font-weight:600;color:#1e293b;cursor:pointer;text-align:left;-webkit-tap-highlight-color:transparent;min-height:68px;display:flex;align-items:center}
.emp-btn:active{background:#f1f5f9}
.emp-btn:disabled{opacity:.45;cursor:not-allowed}
#loading-msg{text-align:center;color:#94a3b8;padding:48px 0;font-size:1.1rem}

/* Screen 2 */
#s-action{align-items:center;text-align:center}
.live-clock{font-size:2.4rem;font-weight:800;color:#1e3a5f;font-variant-numeric:tabular-nums;letter-spacing:.02em;margin-bottom:2px}
.live-date{font-size:1rem;color:#94a3b8;margin-bottom:22px}
.a-greeting{font-size:1.1rem;color:#64748b;font-weight:500}
.a-name{font-size:2rem;font-weight:800;color:#1e293b;margin-bottom:6px;line-height:1.1}
.a-status{font-size:1.05rem;color:#64748b;margin-bottom:28px;min-height:1.5em;line-height:1.4}
#a-btns{display:flex;flex-direction:column;gap:14px;width:100%;max-width:400px}
.action-btn{width:100%;min-height:68px;padding:16px 20px;border-radius:16px;border:none;font-size:1.15rem;font-weight:700;cursor:pointer;-webkit-tap-highlight-color:transparent;line-height:1.35;text-align:center}
.action-btn:disabled{opacity:.45;cursor:not-allowed}
.btn-in{background:#16a34a;color:#fff}
.btn-in:active:not(:disabled){background:#15803d}
.btn-out{background:#dc2626;color:#fff}
.btn-out:active:not(:disabled){background:#b91c1c}
.btn-lunch{background:#d97706;color:#fff}
.btn-lunch:active:not(:disabled){background:#b45309}
.done-today{font-size:1.05rem;font-weight:600;color:#16a34a;padding:22px 24px;background:#f0fdf4;border-radius:14px;border:1.5px solid #bbf7d0;width:100%;text-align:center;line-height:1.5}
.scan-prompt{font-size:1.05rem;font-weight:600;color:#1e3a5f;padding:22px 24px;background:#eff6ff;border-radius:14px;border:1.5px solid #bfdbfe;width:100%;text-align:center;line-height:1.6}
.switch-link{font-size:.9rem;color:#94a3b8;cursor:pointer;padding:28px 16px;margin-top:auto;text-align:center;-webkit-tap-highlight-color:transparent;text-decoration:underline;text-underline-offset:3px}
.switch-link:active{color:#64748b}
.btn-sub{display:block;font-size:.82em;opacity:.8;font-weight:500;margin-top:2px}

/* Full-screen flash confirmation */
#flash{display:none;position:fixed;inset:0;background:#16a34a;flex-direction:column;align-items:center;justify-content:center;text-align:center;padding:40px 32px;z-index:200}
#flash.active{display:flex}
.flash-icon{font-size:88px;margin-bottom:24px}
.flash-msg{font-size:1.9rem;font-weight:800;color:#fff;margin-bottom:10px;line-height:1.2}
.flash-sub{font-size:1.1rem;color:rgba(255,255,255,.88);line-height:1.4}
</style>
</head>
<body>
<header>
  <h1>Willamette Power Roofing</h1>
  <p>Employee Time Clock</p>
</header>
<div id="session-bar"></div>
<div id="error-banner"></div>

<!-- Full-screen confirmation flash -->
<div id="flash">
  <div class="flash-icon" id="flash-icon"></div>
  <div class="flash-msg"  id="flash-msg"></div>
  <div class="flash-sub"  id="flash-sub"></div>
</div>

<!-- Screen 1: Who are you? -->
<div id="s-who" class="screen active">
  <div id="loading-msg">Loading&#8230;</div>
  <h2 id="who-heading" style="display:none">Who are you? / ¿Quién eres?</h2>
  <p class="sub" id="who-sub" style="display:none">Tap your name / Toca tu nombre</p>
  <div id="emp-list"></div>
</div>

<!-- Screen 2: Action -->
<div id="s-action" class="screen">
  <div class="live-clock" id="live-clock"></div>
  <div class="live-date"  id="live-date"></div>
  <div class="a-greeting">Hi / Hola,</div>
  <div class="a-name" id="a-name"></div>
  <div class="a-status" id="a-status"></div>
  <div id="a-btns"></div>
  <div class="switch-link" id="switch-link"></div>
</div>

<script>
var currentEmp=null;

// Scan mode: true on every fresh page load; false after any successful action
sessionStorage.setItem('fromScan','true');

// Live clock — updates every second
function updateClock(){
  var now=new Date();
  var h=now.getHours(),m=now.getMinutes();
  var ap=h>=12?'PM':'AM';var h12=h%12||12;
  var cl=document.getElementById('live-clock');
  var dl=document.getElementById('live-date');
  if(cl)cl.textContent=h12+':'+(m<10?'0':'')+m+' '+ap;
  if(dl){
    var days=['Sunday','Monday','Tuesday','Wednesday','Thursday','Friday','Saturday'];
    var mos=['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'];
    dl.textContent=days[now.getDay()]+', '+mos[now.getMonth()]+' '+now.getDate();
  }
}
updateClock();setInterval(updateClock,10000);

function fmtHHMM(s){
  if(!s)return'';
  var m=s.match(/(\\d{1,2}):(\\d{2})/);
  if(!m)return s;
  var h=parseInt(m[1]),mn=parseInt(m[2]);
  var ap=h>=12?'PM':'AM';h=h%12||12;
  return h+':'+(mn<10?'0':'')+mn+' '+ap;
}

function showError(msg){
  var b=document.getElementById('error-banner');
  b.innerHTML=msg;b.style.display='block';
  setTimeout(function(){b.style.display='none';},6000);
}

function show(id){
  ['s-who','s-action'].forEach(function(s){
    document.getElementById(s).classList.toggle('active',s===id);
  });
}

async function fetchAll(){
  console.log('[tc] fetchAll: starting fetch /employees');
  var ctrl=new AbortController();
  var tid=setTimeout(function(){
    console.log('[tc] fetchAll: 10s timeout reached, aborting');
    ctrl.abort();
  },10000);
  try{
    var r=await fetch('/employees',{signal:ctrl.signal});
    console.log('[tc] fetchAll: response received, status='+r.status);
    var d=await r.json();
    console.log('[tc] fetchAll: parsed JSON, employee count='+(d.employees||[]).length);
    return d.employees||[];
  }finally{
    clearTimeout(tid);
  }
}

function makeBtn(label,sub,cls,handler){
  var b=document.createElement('button');
  b.className='action-btn '+cls;
  b.innerHTML=label+(sub?'<span class="btn-sub">'+sub+'</span>':'');
  b.onclick=handler;
  return b;
}

function showConnectError(){
  var lm=document.getElementById('loading-msg');
  lm.style.display='block';
  lm.innerHTML='Could not connect. Please check your connection and try again.<br>No se pudo conectar. Por favor verifique su conexión e intente de nuevo.<br><button onclick="loadAll()" style="margin-top:16px;padding:14px 32px;background:#1e3a5f;color:#fff;border:none;border-radius:12px;font-size:1rem;font-weight:700;cursor:pointer">Retry / Reintentar</button>';
}

// 10-second global fallback: if still loading, show a message
var _loadingTimer=setTimeout(function(){
  var lm=document.getElementById('loading-msg');
  if(lm&&lm.style.display!=='none'&&!lm.innerHTML.includes('Retry')){
    console.log('[tc] fallback: still loading after 10s');
    lm.innerHTML='Taking too long. Please refresh.<br>Tardando demasiado. Por favor recargue.<br><button onclick="location.reload()" style="margin-top:16px;padding:14px 32px;background:#1e3a5f;color:#fff;border:none;border-radius:12px;font-size:1rem;font-weight:700;cursor:pointer">Refresh / Recargar</button>';
  }
},10000);

async function init(){
  var saved=null;
  try{saved=JSON.parse(localStorage.getItem('tc_employee'))}catch(e){
    console.log('[tc] init: localStorage parse error='+e);
    localStorage.removeItem('tc_employee');
  }
  console.log('[tc] init: localStorage saved='+(saved?saved.name:'none'));
  if(saved&&saved.id){
    show('s-action');
    document.getElementById('a-name').textContent=saved.name;
    document.getElementById('a-status').textContent='Loading…';
    document.getElementById('a-btns').innerHTML='';
    document.getElementById('switch-link').textContent='';
    try{
      var emps=await fetchAll();
      var emp=emps.find(function(e){return e.id===saved.id});
      console.log('[tc] init: saved employee found in list='+(!!emp));
      if(emp){clearTimeout(_loadingTimer);showAction(emp);return;}
      // Stale ID — clear and fall through to employee selection
      console.log('[tc] init: stale employee ID, clearing localStorage');
      localStorage.removeItem('tc_employee');
    }catch(e){
      console.log('[tc] init: fetchAll error='+e);
      show('s-who');
      showConnectError();
      return;
    }
  }
  loadAll();
}

async function loadAll(){
  console.log('[tc] loadAll: starting');
  show('s-who');
  document.getElementById('loading-msg').style.display='block';
  document.getElementById('who-heading').style.display='none';
  document.getElementById('who-sub').style.display='none';
  document.getElementById('emp-list').innerHTML='';
  try{
    var emps=await fetchAll();
    console.log('[tc] loadAll: rendering '+emps.length+' employees');
    clearTimeout(_loadingTimer);
    renderList(emps);
  }catch(e){
    console.log('[tc] loadAll: fetchAll error='+e);
    showConnectError();
  }
}

function renderList(emps){
  document.getElementById('loading-msg').style.display='none';
  var list=document.getElementById('emp-list');
  if(!emps.length){
    list.innerHTML='<p style="text-align:center;color:#94a3b8;padding:40px 0;font-size:1.1rem">No employees set up yet.</p>';
    return;
  }
  document.getElementById('who-heading').style.display='block';
  document.getElementById('who-sub').style.display='block';
  list.innerHTML='';
  emps.forEach(function(emp){
    var btn=document.createElement('button');
    btn.className='emp-btn';
    btn.textContent=emp.name;
    btn.addEventListener('click',function(){
      localStorage.setItem('tc_employee',JSON.stringify({id:emp.id,name:emp.name}));
      showAction(emp);
    });
    list.appendChild(btn);
  });
}

function showAction(emp){
  currentEmp=emp;
  document.getElementById('a-name').textContent=emp.name;
  var statusEl=document.getElementById('a-status');
  var btns=document.getElementById('a-btns');
  btns.innerHTML='';

  var scan=sessionStorage.getItem('fromScan')!=='false';
  var st=emp.status;

  function scanMsg(txt){
    var p=document.createElement('div');
    p.className='scan-prompt';
    p.textContent=txt;
    btns.appendChild(p);
  }

  if(st==='not_clocked_in'){
    // Clock In available in both modes
    statusEl.textContent='Not clocked in yet / No registrado aun';
    btns.appendChild(makeBtn('🏠 Start the Work Day','Registrar Entrada','btn-in',function(){doAction(emp,'clock_in')}));

  } else if(st==='clocked_in'||st==='returned_from_lunch'){
    // returned_from_lunch is treated as clocked_in for display
    statusEl.textContent='Clocked in at / Entrada a las '+fmtHHMM(emp.clock_in_time);
    if(scan){
      btns.appendChild(makeBtn('🍔 Going for Lunch','Salida a Almuerzo','btn-lunch',function(){doAction(emp,'lunch_out')}));
      btns.appendChild(makeBtn('🏁 Done for the Day','Registrar Salida','btn-out',function(){doAction(emp,'clock_out')}));
    } else {
      scanMsg('Scan QR code to continue / Escanee el codigo QR para continuar');
    }

  } else if(st==='on_lunch'){
    statusEl.textContent='On lunch since / En almuerzo desde '+fmtHHMM(emp.lunch_out_time);
    if(scan){
      btns.appendChild(makeBtn('🥪 Back from Lunch','Regreso de Almuerzo','btn-in',function(){doAction(emp,'lunch_in')}));
    } else {
      scanMsg('Scan QR code to continue / Escanee el codigo QR para continuar');
    }
  }

  document.getElementById('switch-link').textContent='Not '+emp.name+'? Switch / Cambiar empleado';
  document.getElementById('switch-link').onclick=function(){
    localStorage.removeItem('tc_employee');
    sessionStorage.setItem('fromScan','true');
    currentEmp=null;
    loadAll();
  };
  show('s-action');
}

function showDoneForToday(){
  if(!currentEmp){loadAll();return;}
  document.getElementById('a-name').textContent=currentEmp.name;
  document.getElementById('a-status').textContent='';
  var btns=document.getElementById('a-btns');
  btns.innerHTML='';
  var d=document.createElement('div');
  d.className='done-today';
  d.innerHTML="🎉 You're clocked out. See you tomorrow! / ¡Hasta manana!";
  btns.appendChild(d);
  document.getElementById('switch-link').textContent='Not '+currentEmp.name+'? Switch / Cambiar empleado';
  document.getElementById('switch-link').onclick=function(){
    localStorage.removeItem('tc_employee');
    sessionStorage.setItem('fromScan','true');
    currentEmp=null;
    loadAll();
  };
  show('s-action');
}

async function doAction(emp,actionType){
  document.querySelectorAll('#a-btns button').forEach(function(b){b.disabled=true;});
  try{
    var r=await fetch('/timeclock/action',{
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body:JSON.stringify({employee_id:emp.id,action:actionType})
    });
    if(!r.ok){
      showAction(emp);
      showError('Something went wrong (error '+r.status+'). Please try again.<br>Algo salió mal. Por favor intente de nuevo.');
      return;
    }
    var d=await r.json();
    showFlash(d);
  }catch(e){
    showAction(emp);
    showError('Could not reach the server. Check your connection.<br>No se pudo conectar. Verifique su conexión.');
  }
}

function showFlash(d){
  // Any completed action switches to stay mode
  sessionStorage.setItem('fromScan','false');

  var map={
    clock_in: {icon:'👋',msg:'Hi, '+d.name+'! / Hola!',         sub:'Clocked in at / Entrada a las '+d.time},
    clock_out:{icon:'✅',msg:'Goodbye! / Hasta luego, '+d.name+'!',sub:'Clocked out at / Salida a las '+d.time},
    lunch_out:{icon:'🍔',msg:'Enjoy lunch! / Buen provecho!',     sub:'Lunch out at / Almuerzo desde '+d.time},
    lunch_in: {icon:'💼',msg:'Welcome back, '+d.name+'!',         sub:'Back at / Regreso a las '+d.time}
  };
  var m=map[d.action]||{icon:'✅',msg:d.name,sub:d.time};
  document.getElementById('flash-icon').textContent=m.icon;
  document.getElementById('flash-msg').textContent=m.msg;
  document.getElementById('flash-sub').textContent=m.sub;
  document.getElementById('flash').className='active';
  setTimeout(async function(){
    document.getElementById('flash').className='';
    if(d.action==='clock_out'){
      showDoneForToday();
      return;
    }
    if(currentEmp){
      try{
        var emps=await fetchAll();
        var emp=emps.find(function(e){return e.id===currentEmp.id});
        if(emp){showAction(emp);return;}
      }catch(e){}
    }
    loadAll();
  },2200);
}

// Session timeout — starts when the page loads, not from QR code timestamp
(function(){
  var t=Math.floor(Date.now()/1000);
  var TIMEOUT=300;
  var bar=document.getElementById('session-bar');
  var expired=false;

  function remaining(){
    return TIMEOUT-(Math.floor(Date.now()/1000)-t);
  }
  function fmt(s){
    var m=Math.ceil(s/60);
    return 'Session expires in / Sesion expira en '+m+' min';
  }
  function disableAll(){
    document.querySelectorAll('.action-btn,.emp-btn').forEach(function(b){b.disabled=true;});
  }
  function expire(){
    if(expired)return;
    expired=true;
    if(bar){
      bar.textContent='Session expired. Please scan the QR code again. / Sesión expirada. Por favor escanee el código QR de nuevo.';
      bar.className='expired';
    }
    disableAll();
    var obs=new MutationObserver(disableAll);
    obs.observe(document.body,{childList:true,subtree:true});
  }
  function tick(){
    var r=remaining();
    if(r<=0){expire();return;}
    if(bar)bar.textContent=fmt(r);
    setTimeout(tick,30000);
  }
  if(!t){expire();}else{tick();}
})();

init();
</script>
</body>
</html>""")


@app.get("/timeclock/qr")
def timeclock_qr():
    base_url = os.getenv("BASE_URL", "https://smartbiz-outreach.onrender.com")
    url = f"{base_url}/timeclock"
    return {"qr_code": _make_qr_b64(url), "url": url}


def _is_blank(val) -> bool:
    """True if a Sheets cell value counts as empty (None, '', or whitespace)."""
    return val is None or (isinstance(val, str) and not val.strip())


@app.post("/timeclock/action")
def timeclock_action(req: TimeclockActionRequest):
    employees = _load_employees()
    emp = next((e for e in employees if e.get("id") == req.employee_id), None)
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found.")

    timesheets = _load_timesheets()
    now_dt     = datetime.now(_TZ)
    today      = now_dt.date().isoformat()
    now_time   = now_dt.strftime("%H:%M")
    now_disp   = _fmt_clock(now_time)

    # All rows for this employee today, in sheet order (oldest first)
    today_rows = [
        t for t in timesheets
        if t.get("employee_id") == emp["id"] and _row_date(t) == today
    ]
    last_row = today_rows[-1] if today_rows else None
    # The open (active) row is the last one only if it has no clock_out
    open_row = last_row if (last_row and _is_blank(last_row.get("clock_out"))) else None

    print(f"[timeclock] emp={emp['name']!r} action={req.action!r} now={now_time!r} open_row={open_row}", flush=True)

    if req.action == "clock_in":
        if open_row:
            raise HTTPException(status_code=400, detail="Already clocked in — clock out first.")
        new_row = {
            "id": str(uuid.uuid4()), "employee_id": emp["id"],
            "employee_name": emp["name"], "clock_in": f"{today} {now_time}",
        }
        _append_timesheet(new_row)
        return {"action": "clock_in", "name": emp["name"], "time": now_disp}

    if not open_row:
        raise HTTPException(status_code=400, detail="No active clock-in found.")
    entry_id = open_row.get("id", "")

    if req.action == "lunch_out":
        _update_cell(entry_id, 5, now_time)  # col 5 = lunch_out
        return {"action": "lunch_out", "name": emp["name"], "time": now_disp}

    if req.action == "lunch_in":
        _update_cell(entry_id, 6, now_time)  # col 6 = lunch_in
        return {"action": "lunch_in", "name": emp["name"], "time": now_disp}

    if req.action == "clock_out":
        ci_full   = (open_row.get("clock_in")  or "").strip()
        lunch_out = (open_row.get("lunch_out") or "").strip()
        lunch_in  = (open_row.get("lunch_in")  or "").strip()
        if lunch_out and not lunch_in:
            notes = "missed lunch clock-in"   # left for lunch, never clocked back in
        elif not lunch_out:
            notes = "no lunch taken on shift"
        else:
            notes = ""
        _update_clock_out(entry_id, now_time, ci_full, lunch_out, "", notes=notes)
        return {"action": "clock_out", "name": emp["name"], "time": now_disp}

    raise HTTPException(status_code=400, detail=f"Unknown action: {req.action!r}")


@app.get("/timesheet")
def get_timesheet():
    entries = _load_timesheets()
    result  = []
    for e in sorted(entries, key=lambda x: (_row_date(x), x.get("employee_name", ""))):
        result.append({**e, "date": _row_date(e), "hours": _calc_hours(e), "lunch_minutes": _calc_lunch(e)})
    return {"entries": result}


@app.get("/timesheet/export")
def export_timesheet():
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    entries     = sorted(_load_timesheets(), key=lambda x: (x.get("employee_name", ""), _row_date(x)))
    wb          = openpyxl.Workbook()
    ws          = wb.active
    ws.title    = "Timesheet"
    hdr_fill    = PatternFill("solid", fgColor="1E40AF")
    yellow_fill = PatternFill("solid", fgColor="FEF9C3")

    headers = ["Employee Name", "Date", "Clock In", "Lunch Out", "Lunch In", "Lunch Time", "Clock Out", "Lunch Hours", "Work Hours", "Notes"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font      = Font(bold=True, color="FFFFFF")
        c.fill      = hdr_fill
        c.alignment = Alignment(horizontal="center")

    for ri, e in enumerate(entries, 2):
        lunch_mins = _calc_lunch(e)
        lunch_str  = ""
        if lunch_mins is not None:
            lh, lm = divmod(lunch_mins, 60)
            lunch_str = f"{lh}h {lm}m" if lh else f"{lm}m"
        row = [
            e.get("employee_name", ""),
            _row_date(e),
            _fmt_clock(e["clock_in"])   if e.get("clock_in")   else "",
            _fmt_clock(e["lunch_out"])  if e.get("lunch_out")  else "",
            _fmt_clock(e["lunch_in"])   if e.get("lunch_in")   else "",
            lunch_str,
            _fmt_clock(e["clock_out"])  if e.get("clock_out")  else "",
            e.get("lunch_hours") or "",
            e.get("work_hours")  or "",
            e.get("notes", "") or "",
        ]
        for col, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=col, value=val)
            if not e.get("clock_out"):
                c.fill = yellow_fill

    for col, w in zip("ABCDEFGHIJ", [22, 12, 12, 11, 11, 11, 12, 11, 11, 18]):
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"timesheet-{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
